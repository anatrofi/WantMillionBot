from aiogram import Bot, types
from aiogram.dispatcher import Dispatcher
from aiogram.utils import executor
from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup

import random
import openpyxl

from messages import MESSAGES
from keybords import BUTTONS

TOKEN = '5477519127:AAHM6onbWedyvvKSBxIGeww0O0viFG1TBgk'

bot = Bot(token=TOKEN)
dp = Dispatcher(bot)
que = openpyxl.load_workbook('вопросы.xlsx')

que_list = [1, 2, 3, 4, 5]
ans_list = [2, 3, 4, 5]


@dp.message_handler(commands=['start'])
async def process_start_command(message: types.Message):
    await message.reply(MESSAGES['start'], reply_markup=BUTTONS['start'])


@dp.callback_query_handler(lambda he: he.data == 'help')
async def helper(callback_query: types.CallbackQuery):
    await bot.send_message(callback_query.from_user.id, MESSAGES['help'], reply_markup=BUTTONS['letsgo'])


@dp.callback_query_handler(lambda first: first.data == 'first')
async def first_question(callback_query: types.CallbackQuery):
    que.active = 0
    sheet = que.active
    qrow = random.sample(que_list, 1)[0]
    quecell = sheet.cell(row=qrow, column=1)
    ans_order = random.sample(ans_list, 4)
    ans_cells = [sheet.cell(row=qrow, column=ans_order[0]), sheet.cell(row=qrow, column=ans_order[1]),
                 sheet.cell(row=qrow, column=ans_order[2]), sheet.cell(row=qrow, column=ans_order[3])]
    ans_buttons = []
    for i in range(4):
        if ans_order[i] == 2:
            ans_buttons.append(InlineKeyboardButton(str(ans_cells[i].value), callback_data='second'))
        else:
            ans_buttons.append(InlineKeyboardButton(str(ans_cells[i].value), callback_data='looser'))
    answers = InlineKeyboardMarkup().add(*ans_buttons)
    await bot.send_message(callback_query.from_user.id, MESSAGES['firstque'] + str(quecell.value), reply_markup=answers)


@dp.callback_query_handler(lambda sec: sec.data == 'second')
async def second_question(callback_query: types.CallbackQuery):
    que.active = 1
    sheet = que.active
    qrow = random.sample(que_list, 1)[0]
    quecell = sheet.cell(row=qrow, column=1)
    ans_order = random.sample(ans_list, 4)
    ans_cells = [sheet.cell(row=qrow, column=ans_order[0]), sheet.cell(row=qrow, column=ans_order[1]),
                 sheet.cell(row=qrow, column=ans_order[2]), sheet.cell(row=qrow, column=ans_order[3])]
    ans_buttons = []
    for i in range(4):
        if ans_order[i] == 2:
            ans_buttons.append(InlineKeyboardButton(str(ans_cells[i].value), callback_data='third'))
        else:
            ans_buttons.append(InlineKeyboardButton(str(ans_cells[i].value), callback_data='looser'))
    answers = InlineKeyboardMarkup().add(*ans_buttons)
    await bot.send_message(callback_query.from_user.id, MESSAGES['secque'] + str(quecell.value), reply_markup=answers)


@dp.callback_query_handler(lambda third: third.data == 'third')
async def third_question(callback_query: types.CallbackQuery):
    que.active = 2
    sheet = que.active
    qrow = random.sample(que_list, 1)[0]
    quecell = sheet.cell(row=qrow, column=1)
    ans_order = random.sample(ans_list, 4)
    ans_cells = [sheet.cell(row=qrow, column=ans_order[0]), sheet.cell(row=qrow, column=ans_order[1]),
                 sheet.cell(row=qrow, column=ans_order[2]), sheet.cell(row=qrow, column=ans_order[3])]
    ans_buttons = []
    for i in range(4):
        if ans_order[i] == 2:
            ans_buttons.append(InlineKeyboardButton(str(ans_cells[i].value), callback_data='four'))
        else:
            ans_buttons.append(InlineKeyboardButton(str(ans_cells[i].value), callback_data='looser'))
    answers = InlineKeyboardMarkup().add(*ans_buttons)
    await bot.send_message(callback_query.from_user.id, MESSAGES['thque'] + str(quecell.value), reply_markup=answers)


@dp.callback_query_handler(lambda four: four.data == 'four')
async def fourth_question(callback_query: types.CallbackQuery):
    que.active = 3
    sheet = que.active
    qrow = random.sample(que_list, 1)[0]
    quecell = sheet.cell(row=qrow, column=1)
    ans_order = random.sample(ans_list, 4)
    ans_cells = [sheet.cell(row=qrow, column=ans_order[0]), sheet.cell(row=qrow, column=ans_order[1]),
                 sheet.cell(row=qrow, column=ans_order[2]), sheet.cell(row=qrow, column=ans_order[3])]
    ans_buttons = []
    for i in range(4):
        if ans_order[i] == 2:
            ans_buttons.append(InlineKeyboardButton(str(ans_cells[i].value), callback_data='five'))
        else:
            ans_buttons.append(InlineKeyboardButton(str(ans_cells[i].value), callback_data='looser'))
    answers = InlineKeyboardMarkup().add(*ans_buttons)
    await bot.send_message(callback_query.from_user.id, MESSAGES['fourque'] + str(quecell.value), reply_markup=answers)


@dp.callback_query_handler(lambda five: five.data == 'five')
async def fifth_question(callback_query: types.CallbackQuery):
    que.active = 4
    sheet = que.active
    qrow = random.sample(que_list, 1)[0]
    quecell = sheet.cell(row=qrow, column=1)
    ans_order = random.sample(ans_list, 4)
    ans_cells = [sheet.cell(row=qrow, column=ans_order[0]), sheet.cell(row=qrow, column=ans_order[1]),
                 sheet.cell(row=qrow, column=ans_order[2]), sheet.cell(row=qrow, column=ans_order[3])]
    ans_buttons = []
    for i in range(4):
        if ans_order[i] == 2:
            ans_buttons.append(InlineKeyboardButton(str(ans_cells[i].value), callback_data='winner'))
        else:
            ans_buttons.append(InlineKeyboardButton(str(ans_cells[i].value), callback_data='looser'))
    answers = InlineKeyboardMarkup().add(*ans_buttons)
    await bot.send_message(callback_query.from_user.id, MESSAGES['fiveque'] + str(quecell.value), reply_markup=answers)


@dp.callback_query_handler(lambda end: end.data == 'looser')
async def loose(callback_query: types.CallbackQuery):
    await bot.send_message(callback_query.from_user.id, MESSAGES['loser'], reply_markup=BUTTONS['choose'])


@dp.callback_query_handler(lambda winn: winn.data == 'winner')
async def win(callback_query: types.CallbackQuery):
    await bot.send_message(callback_query.from_user.id, MESSAGES['ifwin'], reply_markup=BUTTONS['again'])


if __name__ == '__main__':
    executor.start_polling(dp)
